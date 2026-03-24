"""
Verify calendar rotation in the database matches the original Excel spreadsheet.
Compares years 2026-2036 by default (configurable).

Checks:
  1. Rotation dates: Does each owner have the same week_start dates in the DB
     as in the Excel "Calendar to 2100" sheet?
  2. Trade detail: Does each (owner, week_start) in the DB trade_detail table
     match the Excel "TradeDetail" sheet?

Run from project root:
    python -m snowbound.scripts.verify_calendar
    python -m snowbound.scripts.verify_calendar --years 2026-2050
"""
import sys
import os
import re
import argparse

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import openpyxl
from snowbound import create_app, db
from snowbound.models import Owner, Calendar, TradeDetail

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX_PATH = os.path.join(PROJECT_ROOT, "GAS-code", "SnowBoundersCalendarApp.xlsx")

# Map full_name prefixes in the Excel owner_info column to short_names in the DB
OWNER_INFO_TO_SHORT = {
    "Jim & Janice Boone": "Boone",
    "Larry & Maureen Kamons": "Kamons",
    "Linda Mitchell": "Mitchell",
    "Pam Loyle": "Loyle",
    "Dave & Bob Zerfas": "Zerfas",
    "Dave & Sid Sproul": "Sproul",
    "Stan & Joni Miller": "Miller",
    "Dave & Sharon Stalker": "Stalker",
    "Brad & Cathrine Smith": "Smith",
}


def extract_owner_short_name(owner_info_cell):
    """Extract the short_name from the multiline owner_info cell in Excel."""
    if not owner_info_cell:
        return None
    first_line = str(owner_info_cell).split("\n")[0].strip()
    for full, short in OWNER_INFO_TO_SHORT.items():
        if first_line.startswith(full):
            return short
    return first_line  # fallback


def extract_date_from_week_cell(cell_value):
    """Extract just the MM/DD/YYYY date from a week cell that may have comments appended."""
    if not cell_value:
        return None
    s = str(cell_value).strip()
    # Match MM/DD/YYYY at the start
    m = re.match(r"(\d{2}/\d{2}/\d{4})", s)
    return m.group(1) if m else None


def read_excel_rotation(xlsx_path, start_year, end_year):
    """
    Read "Calendar to 2100" sheet.
    Returns dict: {year: [(short_name, [week1, week2, ...]), ...]}
    Loyle appears twice per year (two rows).
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Calendar to 2100"]

    result = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        year_val = row[0]
        if year_val is None:
            continue
        try:
            year = int(year_val)
        except (ValueError, TypeError):
            continue

        if year < start_year or year > end_year:
            continue

        owner_short = extract_owner_short_name(row[1])
        weeks = []
        for col_idx in range(2, 7):  # columns C-G = weeks 1-5
            date_str = extract_date_from_week_cell(row[col_idx])
            if date_str:
                weeks.append(date_str)

        if year not in result:
            result[year] = []
        result[year].append((owner_short, weeks))

    wb.close()
    return result


def read_excel_trade_detail(xlsx_path, start_year, end_year):
    """
    Read "TradeDetail" sheet.
    Returns dict: {(year, owner_name, week_start): {is_traded, calculated_owner, comment}}
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["TradeDetail"]

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        year_val, name, week_start, traded = row[0], row[1], row[2], row[3]
        comment = row[7] if len(row) > 7 else None
        calc_owner = row[9] if len(row) > 9 else None

        if year_val is None or name is None or week_start is None:
            continue

        try:
            year = int(year_val)
        except (ValueError, TypeError):
            continue

        if year < start_year or year > end_year:
            continue

        name = str(name).strip()
        week_start = str(week_start).strip()
        is_traded = bool(traded and str(traded).strip().upper() in ("Y", "YES", "TRUE", "1"))
        calc_owner_str = str(calc_owner).strip() if calc_owner else None
        comment_str = str(comment).strip() if comment else None

        result[(year, name, week_start)] = {
            "is_traded": is_traded,
            "calculated_owner": calc_owner_str,
            "comment": comment_str,
        }

    wb.close()
    return result


def read_db_rotation(start_year, end_year):
    """
    Read calendar table from DB.
    Returns dict: {year: [(short_name, week_start, week_number), ...]}
    """
    result = {}
    rows = (Calendar.query
            .join(Owner)
            .filter(Calendar.year >= start_year, Calendar.year <= end_year)
            .order_by(Calendar.year, Calendar.week_start)
            .all())
    for r in rows:
        year = r.year
        if year not in result:
            result[year] = []
        result[year].append((r.owner.short_name, r.week_start, r.week_number))
    return result


def read_db_trade_detail(start_year, end_year):
    """
    Read trade_detail table from DB.
    Returns dict: {(year, short_name, week_start): {is_traded, calculated_owner, comment}}
    """
    result = {}
    rows = (TradeDetail.query
            .join(Owner, TradeDetail.owner_id == Owner.id)
            .filter(TradeDetail.year >= start_year, TradeDetail.year <= end_year)
            .all())
    for r in rows:
        key = (r.year, r.owner.short_name, r.week_start)
        result[key] = {
            "is_traded": bool(r.is_traded),
            "calculated_owner": r.calculated_owner,
            "comment": r.comment,
        }
    return result


def compare_rotation(excel_rot, db_rot, start_year, end_year):
    """Compare rotation dates between Excel and DB."""
    errors = []
    for year in range(start_year, end_year + 1):
        excel_rows = excel_rot.get(year, [])
        db_rows_raw = db_rot.get(year, [])

        # Build DB data: group by owner, preserving order of appearance
        # (Loyle has two groups of weeks)
        from collections import OrderedDict
        db_owner_weeks = OrderedDict()
        for short_name, week_start, week_num in db_rows_raw:
            if short_name not in db_owner_weeks:
                db_owner_weeks[short_name] = []
            db_owner_weeks[short_name].append(week_start)

        # Excel data: each row is (short_name, [weeks])
        # Loyle appears twice, so we need to handle that
        # Flatten Excel into owner -> all weeks (in row order)
        excel_owner_weeks = OrderedDict()
        for short_name, weeks in excel_rows:
            if short_name not in excel_owner_weeks:
                excel_owner_weeks[short_name] = []
            excel_owner_weeks[short_name].extend(weeks)

        # Compare owners present
        excel_owners = set(excel_owner_weeks.keys())
        db_owners = set(db_owner_weeks.keys())

        if excel_owners != db_owners:
            missing_in_db = excel_owners - db_owners
            extra_in_db = db_owners - excel_owners
            if missing_in_db:
                errors.append(f"  {year}: Owners in Excel but not DB: {missing_in_db}")
            if extra_in_db:
                errors.append(f"  {year}: Owners in DB but not Excel: {extra_in_db}")

        # Compare week dates per owner
        for owner in sorted(excel_owners & db_owners):
            excel_weeks = excel_owner_weeks.get(owner, [])
            db_weeks = db_owner_weeks.get(owner, [])

            if excel_weeks != db_weeks:
                errors.append(f"  {year} {owner}: MISMATCH")
                errors.append(f"    Excel ({len(excel_weeks)}): {excel_weeks}")
                errors.append(f"    DB    ({len(db_weeks)}):    {db_weeks}")

    return errors


def compare_trade_detail(excel_td, db_td):
    """Compare trade detail between Excel and DB."""
    errors = []

    excel_keys = set(excel_td.keys())
    db_keys = set(db_td.keys())

    missing_in_db = excel_keys - db_keys
    extra_in_db = db_keys - excel_keys

    if missing_in_db:
        errors.append(f"  Trade rows in Excel but not DB: {len(missing_in_db)}")
        for k in sorted(missing_in_db)[:10]:
            errors.append(f"    {k}")
        if len(missing_in_db) > 10:
            errors.append(f"    ... and {len(missing_in_db) - 10} more")

    if extra_in_db:
        errors.append(f"  Trade rows in DB but not Excel: {len(extra_in_db)}")
        for k in sorted(extra_in_db)[:10]:
            errors.append(f"    {k}")
        if len(extra_in_db) > 10:
            errors.append(f"    ... and {len(extra_in_db) - 10} more")

    # Compare shared rows
    trade_mismatches = 0
    for key in sorted(excel_keys & db_keys):
        ex = excel_td[key]
        db = db_td[key]
        diffs = []
        if ex["is_traded"] != db["is_traded"]:
            diffs.append(f"is_traded: Excel={ex['is_traded']} DB={db['is_traded']}")
        # Normalize None vs "None" for comparison
        ex_co = ex["calculated_owner"] if ex["calculated_owner"] not in (None, "None") else None
        db_co = db["calculated_owner"] if db["calculated_owner"] not in (None, "None") else None
        if ex_co != db_co:
            diffs.append(f"calc_owner: Excel={ex_co!r} DB={db_co!r}")
        if diffs:
            trade_mismatches += 1
            if trade_mismatches <= 20:
                errors.append(f"  {key}: {'; '.join(diffs)}")

    if trade_mismatches > 20:
        errors.append(f"  ... and {trade_mismatches - 20} more trade mismatches")
    elif trade_mismatches > 0:
        errors.append(f"  Total trade field mismatches: {trade_mismatches}")

    return errors


def run():
    parser = argparse.ArgumentParser(description="Verify calendar against Excel")
    parser.add_argument("--years", default="2026-2036",
                        help="Year range, e.g. 2026-2036 (default)")
    args = parser.parse_args()

    start_year, end_year = map(int, args.years.split("-"))

    print(f"=== Calendar Verification: {start_year}-{end_year} ===")
    print(f"Excel: {XLSX_PATH}")
    print()

    # Read Excel data
    print("Reading Excel 'Calendar to 2100' sheet...")
    excel_rot = read_excel_rotation(XLSX_PATH, start_year, end_year)
    print(f"  Found rotation data for {len(excel_rot)} years")

    print("Reading Excel 'TradeDetail' sheet...")
    excel_td = read_excel_trade_detail(XLSX_PATH, start_year, end_year)
    print(f"  Found {len(excel_td)} trade detail rows")

    # Read DB data
    app = create_app()
    with app.app_context():
        print("Reading DB calendar table...")
        db_rot = read_db_rotation(start_year, end_year)
        print(f"  Found rotation data for {len(db_rot)} years")

        print("Reading DB trade_detail table...")
        db_td = read_db_trade_detail(start_year, end_year)
        print(f"  Found {len(db_td)} trade detail rows")

        print()

        # === Check 1: Rotation ===
        print("=" * 60)
        print("CHECK 1: Rotation (Calendar to 2100 vs calendar table)")
        print("=" * 60)
        rot_errors = compare_rotation(excel_rot, db_rot, start_year, end_year)
        if rot_errors:
            print("FAILURES:")
            for e in rot_errors:
                print(e)
        else:
            print("PASS — All rotation dates match!")

        print()

        # === Check 2: Trade Detail ===
        print("=" * 60)
        print("CHECK 2: Trade Detail (TradeDetail sheet vs trade_detail table)")
        print("=" * 60)
        td_errors = compare_trade_detail(excel_td, db_td)
        if td_errors:
            print("DIFFERENCES:")
            for e in td_errors:
                print(e)
        else:
            print("PASS — All trade detail rows match!")

        print()

        # === Summary ===
        total_errors = len(rot_errors) + len(td_errors)
        if total_errors == 0:
            print("ALL CHECKS PASSED")
        else:
            print(f"ISSUES FOUND: {len(rot_errors)} rotation issues, {len(td_errors)} trade detail issues")


if __name__ == "__main__":
    run()
